import pandas as pd
import os

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from typing import Optional

def get_display_width(text):
        width = 0
        for ch in str(text):
            if ch in " .,;-_/":
                width += 0
            elif ord(ch) > 255: # 中文或全形
                width += 4
            else:
                width += 1.5
        return width

# Step 1
def export_universe_1(df: pd.DataFrame, before_format: Optional[str] = None, after_format: Optional[str] = None):
    # Export a list of ticker symbols to Excel with auto column width.
    if before_format is None:
        before_format = os.path.abspath("data_clean/universe_step1.xlsx")
    if after_format is None:
        after_format = os.path.abspath("data_clean/universe_step1_formatted.xlsx")

    # Ensure the folder exists
    os.makedirs(os.path.dirname(before_format), exist_ok=True)

    # If the old file exists, combine them
    if os.path.exists(before_format):
        existing_df = pd.read_excel(before_format)
        combined_df = pd.concat([existing_df, df], ignore_index=True)
        # 去重，以 Symbol 為基準，保留最後一次
        combined_df = combined_df.drop_duplicates(subset="Symbol", keep="last")
    else:
        combined_df = df.copy()

    # Export to Excel
    combined_df.to_excel(before_format, index=False)

    # Load workbook for formatting
    wb = load_workbook(before_format)
    ws = wb.active
    if ws is None:
        raise ValueError("No active worksheet found in workbook")

    # Auto-adjust column width
    for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
        col_letter = get_column_letter(col_cells[0].column or 1)
        max_length = 0
        for cell in col_cells:
            if cell.value is not None:
                length = get_display_width(cell.value)
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = max_length

    # Centre all cells
    center = Alignment(horizontal="center", vertical="center")
    times_new_roman = Font(name="Times New Roman")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center
            cell.font = times_new_roman

    # Save and open
    wb.save(after_format)
    os.startfile(after_format)

    print(f"Universe exported: {after_format}")



# Step 2
def export_universe_2(df: pd.DataFrame, before_format: Optional[str] = None, after_format: Optional[str] = None):
    # Export a list of ticker symbols to Excel with auto column width.
    if before_format is None:
        before_format = os.path.abspath("data_metrics/universe_step2.xlsx")
    if after_format is None:
        after_format = os.path.abspath("data_metrics/universe_step2_formatted.xlsx")

    # Ensure the folder exists
    os.makedirs(os.path.dirname(before_format), exist_ok=True)

    # If the old file exists, combine them
    if os.path.exists(before_format):
        existing_df = pd.read_excel(before_format)
        combined_df = pd.concat([existing_df, df], ignore_index=True)
        # 去重，以 Symbol 為基準，保留最後一次
        combined_df = combined_df.drop_duplicates(subset=["Symbol", "date"], keep="last")
    else:
        combined_df = df.copy()

    # Export to Excel
    combined_df.to_excel(before_format, index=False)

    # Load workbook for formatting
    wb = load_workbook(before_format)
    ws = wb.active
    if ws is None:
        raise ValueError("No active worksheet found in workbook")

    # Auto-adjust column width
    for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
        col_letter = get_column_letter(col_cells[0].column or 1)
        max_length = 0
        for cell in col_cells:
            if cell.value is not None:
                length = get_display_width(cell.value)
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = max_length

    # Centre all cells
    center = Alignment(horizontal="center", vertical="center")
    times_new_roman = Font(name="Times New Roman")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center
            cell.font = times_new_roman

    # Save and open
    wb.save(after_format)
    os.startfile(after_format)

    print(f"Universe exported: {after_format}")



# Step 3
def export_universe_3(df: pd.DataFrame, before_format: Optional[str] = None, after_format: Optional[str] = None):
    # Export a list of ticker symbols to Excel with auto column width.
    if before_format is None:
        before_format = os.path.abspath("")
    if after_format is None:
        after_format = os.path.abspath("")

    # Ensure the folder exists
    os.makedirs(os.path.dirname(before_format), exist_ok=True)

    # If the old file exists, combine them
    if os.path.exists(before_format):
        existing_df = pd.read_excel(before_format)
        combined_df = pd.concat([existing_df, df], ignore_index=True)
        # 去重，以 Symbol 為基準，保留最後一次
        combined_df = combined_df.drop_duplicates(subset=["Symbol", "date"], keep="last")
    else:
        combined_df = df.copy()

    # Export to Excel
    combined_df.to_excel(before_format, index=False)

    # Load workbook for formatting
    wb = load_workbook(before_format)
    ws = wb.active
    if ws is None:
        raise ValueError("No active worksheet found in workbook")

    # Auto-adjust column width
    for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
        col_letter = get_column_letter(col_cells[0].column or 1)
        max_length = 0
        for cell in col_cells:
            if cell.value is not None:
                length = get_display_width(cell.value)
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = max_length

    # Centre all cells
    center = Alignment(horizontal="center", vertical="center")
    times_new_roman = Font(name="Times New Roman")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center
            cell.font = times_new_roman

    # Save and open
    wb.save(after_format)
    os.startfile(after_format)

    print(f"Universe exported: {after_format}")